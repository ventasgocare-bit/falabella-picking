#!/usr/bin/env python3
"""
falabella_picking.py

Procesa manifiestos PDF de Falabella Seller Center y genera:
  - LISTA_PICKING.xlsx
  - LISTA_PICKING.pdf

Los archivos se crean en la misma carpeta del PDF de entrada.

Uso:
    python falabella_picking.py <ruta_al_manifiesto.pdf>

Dependencias:
    pip install pdfplumber pandas openpyxl reportlab
"""

import sys
import os
import subprocess
from pathlib import Path
from datetime import datetime

import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER


# ---------------------------------------------------------------------------
# Configuración de salida
# ---------------------------------------------------------------------------

OUTPUT_DIR      = Path.home() / "Documents" / "Salidas Falabella"
OUTPUT_BASENAME = "LISTA_PICKING"

# Nombres internos de columnas
COL_ORDEN       = "N_ORDEN"
COL_SEGUIMIENTO = "N_SEGUIMIENTO"
COL_CANTIDAD    = "CANTIDAD"
COL_ULTIMOS4    = "ULTIMOS_4"

# -- Futura columna de ubicación de bodega ----------------------------------
# Para activar:
#   1. Implementar get_ubicacion() con el mapeo real.
#   2. Cambiar UBICACION_ENABLED = True.
#   3. Descomentar los bloques marcados con "# [UBICACION]" en cada función.
COL_UBICACION     = "UBICACION"
UBICACION_ENABLED = False


# ---------------------------------------------------------------------------
# Stub de ubicación (listo para implementar)
# ---------------------------------------------------------------------------

def get_ubicacion(ultimos_4: int) -> str:
    """
    Retorna la ubicación de bodega para un código ULTIMOS_4.
    Implementar con rangos o lookup table según el layout del almacén.

    Ejemplo de implementación futura:
        if ultimos_4 < 1000: return "Pasillo A"
        if ultimos_4 < 5000: return "Pasillo B"
        return "Pasillo C"
    """
    return ""


# ---------------------------------------------------------------------------
# 1. Extracción del PDF
# ---------------------------------------------------------------------------

def extract_rows(pdf_path: str) -> list[dict]:
    """
    Lee todas las páginas del manifiesto y devuelve las filas de datos.
    Descarta encabezados repetidos y la fila de totales.
    """
    SKIP_KEYWORDS   = {"total de paquetes"}
    HEADER_KEYWORDS = {"nº orden", "numero de seguimiento", "número de seguimiento", "cantidad"}

    rows = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for raw_row in table:
                    if not raw_row or not any(raw_row):
                        continue

                    cells = [str(c).strip() if c else "" for c in raw_row]
                    row_text = " ".join(cells).lower()

                    # Saltar encabezados y fila de total
                    if any(kw in row_text for kw in SKIP_KEYWORDS | HEADER_KEYWORDS):
                        continue

                    # Solo filas cuya primera celda sea un número de orden
                    if len(cells) >= 3 and cells[0].isdigit():
                        rows.append({
                            COL_ORDEN:       cells[0],
                            COL_SEGUIMIENTO: cells[1],
                            COL_CANTIDAD:    cells[2],
                        })

    return rows


# ---------------------------------------------------------------------------
# 2. Construcción del DataFrame
# ---------------------------------------------------------------------------

def build_dataframe(rows: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(rows)

    df[COL_ULTIMOS4] = df[COL_ORDEN].str[-4:].astype(int)
    df[COL_CANTIDAD] = pd.to_numeric(df[COL_CANTIDAD], errors="coerce").fillna(1).astype(int)

    # [UBICACION] Descomentar para activar la columna de bodega:
    # if UBICACION_ENABLED:
    #     df[COL_UBICACION] = df[COL_ULTIMOS4].apply(get_ubicacion)

    df = df.sort_values(COL_ULTIMOS4).reset_index(drop=True)
    return df


# ---------------------------------------------------------------------------
# 3. Generador Excel
# ---------------------------------------------------------------------------

def generate_xlsx(df: pd.DataFrame, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Picking"

    # Paleta
    COLOR_HEADER_BG = "1F3864"
    COLOR_ALT_BG    = "EBF3FB"

    font_header  = Font(name="Calibri", bold=True,  size=11, color="FFFFFF")
    font_bold    = Font(name="Calibri", bold=True,  size=12)
    font_normal  = Font(name="Calibri", bold=False, size=11)
    fill_header  = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    fill_alt     = PatternFill("solid", fgColor=COLOR_ALT_BG)
    align_center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers     = ["#", "ÚLTIMOS 4", "Nº ORDEN", "Nº SEGUIMIENTO", "CANTIDAD"]
    col_widths  = [6,   14,          16,          24,               12]

    # [UBICACION] Descomentar:
    # if UBICACION_ENABLED:
    #     headers.append("UBICACIÓN")
    #     col_widths.append(18)

    # Fila de encabezado
    for col_i, label in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i, value=label)
        cell.font      = font_header
        cell.fill      = fill_header
        cell.alignment = align_center
        cell.border    = border

    ws.row_dimensions[1].height = 22

    # Filas de datos
    for row_i, (_, row) in enumerate(df.iterrows(), 2):
        is_alt = (row_i % 2 == 0)
        fill   = fill_alt if is_alt else None

        values = [
            row_i - 1,
            f"{row[COL_ULTIMOS4]:04d}",
            row[COL_ORDEN],
            row[COL_SEGUIMIENTO],
            int(row[COL_CANTIDAD]),
        ]
        # [UBICACION] Descomentar:
        # if UBICACION_ENABLED:
        #     values.append(row[COL_UBICACION])

        for col_i, value in enumerate(values, 1):
            cell           = ws.cell(row=row_i, column=col_i, value=value)
            cell.alignment = align_center
            cell.border    = border
            cell.font      = font_bold if col_i == 2 else font_normal
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_i].height = 18

    for col_i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, col_i).column_letter].width = width

    wb.save(output_path)


# ---------------------------------------------------------------------------
# 4. Generador PDF
# ---------------------------------------------------------------------------

def generate_pdf(df: pd.DataFrame, output_path: str, source_name: str) -> None:
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()

    style_title = ParagraphStyle(
        "PickingTitle",
        fontName="Helvetica-Bold",
        fontSize=22,
        alignment=TA_CENTER,
        spaceAfter=3,
    )
    style_subtitle = ParagraphStyle(
        "PickingSubtitle",
        fontName="Helvetica",
        fontSize=8,
        textColor=colors.HexColor("#666666"),
        alignment=TA_CENTER,
        spaceAfter=6,
    )
    style_th = ParagraphStyle(
        "TH",
        fontName="Helvetica-Bold",
        fontSize=10,
        alignment=TA_CENTER,
        leading=13,
    )
    style_td = ParagraphStyle(
        "TD",
        fontName="Helvetica",
        fontSize=10,
        alignment=TA_CENTER,
        leading=13,
    )
    # ÚLTIMOS 4: negrita y fuente más grande que el resto
    style_td_u4 = ParagraphStyle(
        "TD_U4",
        fontName="Helvetica-Bold",
        fontSize=15,
        alignment=TA_CENTER,
        leading=18,
    )

    story = []

    # Título
    story.append(Paragraph("LISTA DE PICKING", style_title))

    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    story.append(Paragraph(
        f"Generado: {fecha}&nbsp;&nbsp;|&nbsp;&nbsp;"
        f"Origen: {source_name}&nbsp;&nbsp;|&nbsp;&nbsp;"
        f"Total órdenes: {len(df)}",
        style_subtitle,
    ))
    story.append(Spacer(1, 4 * mm))

    # Columnas y anchos en mm (total ≈ 180 mm con márgenes de 15 mm)
    col_headers_labels = ["#", "ÚLTIMOS 4", "Nº ORDEN", "Nº SEGUIMIENTO", "CANT."]
    col_widths_mm      = [10,  24,           32,          88,               18]

    # [UBICACION] Descomentar:
    # if UBICACION_ENABLED:
    #     col_headers_labels.append("UBICACIÓN")
    #     col_widths_mm[-1] = 14   # reducir columna anterior para hacer espacio
    #     col_widths_mm.append(32)

    col_widths_pt = [w * mm for w in col_widths_mm]

    # Encabezado de tabla
    table_data = [[Paragraph(h, style_th) for h in col_headers_labels]]

    # Filas de datos
    for i, (_, row) in enumerate(df.iterrows(), 1):
        cells = [
            Paragraph(str(i),                   style_td),
            Paragraph(f"{row[COL_ULTIMOS4]:04d}", style_td_u4),   # negrita + grande
            Paragraph(str(row[COL_ORDEN]),       style_td),
            Paragraph(str(row[COL_SEGUIMIENTO]), style_td),
            Paragraph(str(int(row[COL_CANTIDAD])), style_td),
        ]
        # [UBICACION] Descomentar:
        # if UBICACION_ENABLED:
        #     cells.append(Paragraph(str(row[COL_UBICACION]), style_td))
        table_data.append(cells)

    table = Table(table_data, colWidths=col_widths_pt, repeatRows=1)

    # Construir estilos base
    COLOR_HEADER = colors.HexColor("#1F3864")
    COLOR_ALT    = colors.HexColor("#EBF3FB")
    COLOR_BORDER = colors.HexColor("#CCCCCC")

    ts = [
        # Encabezado
        ("BACKGROUND",    (0, 0), (-1, 0), COLOR_HEADER),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("LINEBELOW",     (0, 0), (-1, 0), 1.5, colors.HexColor("#142952")),

        # Bordes generales
        ("GRID",          (0, 0), (-1, -1), 0.5, COLOR_BORDER),

        # Alineación y padding
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
    ]

    # Filas alternas
    for row_i in range(1, len(df) + 1):
        if row_i % 2 == 0:
            ts.append(("BACKGROUND", (0, row_i), (-1, row_i), COLOR_ALT))

    table.setStyle(TableStyle(ts))
    story.append(table)

    doc.build(story)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    if len(sys.argv) < 2:
        print("Uso: python falabella_picking.py <ruta_al_manifiesto.pdf>")
        sys.exit(1)

    pdf_path = sys.argv[1]

    if not os.path.isfile(pdf_path):
        print(f"Error: no se encontró '{pdf_path}'")
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp   = datetime.now().strftime("%Y-%m-%d_%H-%M")
    out_xlsx    = OUTPUT_DIR / f"{OUTPUT_BASENAME}_{timestamp}.xlsx"
    out_pdf     = OUTPUT_DIR / f"{OUTPUT_BASENAME}_{timestamp}.pdf"
    source_name = Path(pdf_path).name

    print(f"[1/4] Leyendo: {pdf_path}")
    rows = extract_rows(pdf_path)
    print(f"      {len(rows)} filas extraídas")

    if not rows:
        print("Error: no se encontraron filas de datos en el PDF.")
        sys.exit(1)

    print("[2/4] Procesando datos...")
    df = build_dataframe(rows)
    print(f"      {len(df)} órdenes | ordenadas por ULTIMOS_4")

    print(f"[3/4] Generando Excel  → {out_xlsx}")
    generate_xlsx(df, str(out_xlsx))

    print(f"[4/4] Generando PDF    → {out_pdf}")
    generate_pdf(df, str(out_pdf), source_name)

    print("\nListo.")
    print(f"  {out_xlsx}")
    print(f"  {out_pdf}")

    try:
        subprocess.run(["open", str(out_pdf)])
    except Exception:
        pass


if __name__ == "__main__":
    main()
